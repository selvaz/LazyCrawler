# -*- coding: utf-8 -*-
"""
lazycrawler.http
================
HTTP client with retry/backoff + URL utilities (normalization, hashing,
domain blacklist).

Hash functions used by the 3-level dedup (see db.py):
  - url_hash(url)        -> sha256(normalize_url(url))   [level 1 dedup, URL]
  - content_hash(text)   -> sha256(normalize(text))      [level 2 dedup, content]
"""

from __future__ import annotations

import hashlib
import ipaddress
import re
import socket
import threading
import time
import weakref
from dataclasses import dataclass
from typing import Dict, List, Optional, Set
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse
from urllib.robotparser import RobotFileParser

import requests

from ._log import log
from .config import HTTPConfig

# =============================================================================
# URL CONSTANTS
# =============================================================================

# Default link-exclusion fragments. Intentionally conservative for a *generic*
# crawler: it drops auth/commerce/tracking/social noise but NOT content-y paths
# like /about, /contact, /tag/, /category/ or /author/ (those are often real
# content). Override via CrawlerConfig.exclude_patterns.
DEFAULT_EXCLUDE_PATTERNS: List[str] = [
    r"unsubscribe",
    r"manage.prefer",
    r"opt.out",
    r"privacy.polic",
    r"terms.of",
    r"facebook\.com",
    r"twitter\.com",
    r"x\.com/",
    r"instagram\.com",
    r"youtube\.com",
    r"substack\.com/subscribe",
    r"mailto:",
    r"tel:",
    r"/login",
    r"/signin",
    r"/register",
    r"/signup",
    r"/cart",
    r"/checkout",
    r"/account",
    r"/search\?",
]


def compile_exclude(patterns: Optional[List[str]] = None) -> "re.Pattern[str]":
    """Compile a link-exclusion regex from fragments (None = built-in default)."""
    frags = DEFAULT_EXCLUDE_PATTERNS if patterns is None else list(patterns)
    if not frags:
        # Match nothing (exclude disabled).
        return re.compile(r"(?!x)x")
    return re.compile("|".join(frags), re.IGNORECASE)


_EXCLUDE_RE = compile_exclude()

_TRACKING_PARAMS: Set[str] = {
    "utm_source",
    "utm_medium",
    "utm_campaign",
    "utm_term",
    "utm_content",
    "utm_id",
    "utm_name",
    "utm_reader",
    "utm_viz_id",
    "gclid",
    "fbclid",
    "mc_cid",
    "mc_eid",
    "cmpid",
    "icid",
    "iid",
    "ref",
    "referrer",
    "source",
    "ns_campaign",
    "ns_mchannel",
    "ns_source",
}


# =============================================================================
# URL UTILITIES
# =============================================================================


def get_base_domain(url: str) -> str:
    """Lowercase domain (netloc) from a URL."""
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ""


def get_hostname(url: str) -> str:
    """Lowercase hostname (no port, no userinfo) from a URL or bare host."""
    try:
        p = urlparse(url)
        host = p.hostname
        if host:
            return host.lower()
    except Exception:
        pass
    # Bare host (no scheme): urlparse puts it in ``path``; strip any port/userinfo.
    bare = (url or "").strip().lower()
    if "://" not in bare and bare:
        bare = bare.split("/")[0].split("@")[-1].split(":")[0]
        return bare
    return ""


# Minimal bundled multi-part public-suffix set for the dependency-free fallback
# (used only when ``tldextract`` is not installed). Not exhaustive; ``tldextract``
# is the accurate path (install the ``domains`` extra).
_MULTI_PART_SUFFIXES = frozenset(
    {
        "co.uk",
        "org.uk",
        "gov.uk",
        "ac.uk",
        "me.uk",
        "com.au",
        "net.au",
        "org.au",
        "gov.au",
        "edu.au",
        "co.jp",
        "or.jp",
        "ne.jp",
        "ac.jp",
        "go.jp",
        "co.nz",
        "org.nz",
        "govt.nz",
        "co.za",
        "org.za",
        "com.br",
        "com.cn",
        "com.sg",
        "com.hk",
        "co.in",
        "co.kr",
    }
)

# Cached tldextract extractor (built once, pinned to the bundled suffix snapshot so
# it never refreshes over the network). ``False`` means "not yet attempted".
_TLD_EXTRACT = False


def _tld_extractor():
    """Return a cached, network-free ``tldextract`` extractor, or None if absent."""
    global _TLD_EXTRACT
    if _TLD_EXTRACT is False:
        try:
            import tldextract  # type: ignore

            # suffix_list_urls=() disables network refresh -> uses bundled snapshot.
            _TLD_EXTRACT = tldextract.TLDExtract(suffix_list_urls=())
        except Exception:
            _TLD_EXTRACT = None
    return _TLD_EXTRACT


def registrable_domain(host_or_url: str) -> str:
    """
    Registrable domain (eTLD+1) of a URL or bare host, lowercased.

    Examples: ``news.example.com`` -> ``example.com``; ``x.bbc.co.uk`` ->
    ``bbc.co.uk``. Ports and userinfo are stripped first. Uses ``tldextract``
    (bundled public-suffix snapshot) when available; otherwise falls back to a
    small built-in multi-part-suffix heuristic. Returns "" if no host.
    """
    host = get_hostname(host_or_url)
    if not host:
        return ""
    # IP literals have no registrable domain - compare them as-is.
    try:
        ipaddress.ip_address(host)
        return host
    except ValueError:
        pass
    ext = _tld_extractor()
    if ext is not None:
        try:
            r = ext(host)
            if r.domain and r.suffix:
                return f"{r.domain}.{r.suffix}".lower()
            # No known public suffix (e.g. test/fake TLD) -> heuristic below.
        except Exception:
            pass
    return _registrable_heuristic(host)


def _registrable_heuristic(host: str) -> str:
    """Dependency-free eTLD+1 guess using a small bundled multi-part-suffix set."""
    labels = host.split(".")
    if len(labels) <= 2:
        return host
    last2 = ".".join(labels[-2:])
    last3 = ".".join(labels[-3:])
    if last2 in _MULTI_PART_SUFFIXES:
        return ".".join(labels[-3:])
    if last3 in _MULTI_PART_SUFFIXES:
        return ".".join(labels[-4:]) if len(labels) >= 4 else host
    return last2


def same_site(host_a: str, host_b: str) -> bool:
    """True if two hosts/URLs share the same registrable domain (eTLD+1)."""
    ra = registrable_domain(host_a)
    rb = registrable_domain(host_b)
    return bool(ra) and ra == rb


_METADATA_HOSTS = {"metadata.google.internal", "metadata"}


def is_blocked_address(url: str) -> bool:
    """
    SSRF guard: True if ``url`` targets a non-public address.

    Blocks loopback / link-local / private (RFC-1918) / reserved / multicast /
    unspecified IPs, ``localhost`` / ``*.local`` hosts, and known cloud metadata
    endpoints (e.g. 169.254.169.254 via link-local). The host is resolved with
    ``socket.getaddrinfo`` and every returned address is checked; a resolution
    failure or unparseable host is treated as blocked (fail-closed).

    Intended for the agent/tool path where an LLM may pass arbitrary URLs. Off by
    default for the library (``HTTPConfig.block_private_addresses``).

    .. warning::
       **Best-effort guard, not network isolation.** This validates the IPs the
       host resolves to *at check time*, but the connection is established
       separately (by ``requests``/``aiohttp``), which re-resolves the host. A
       hostile DNS server can therefore return a public IP during validation and a
       private/loopback IP at connect time (DNS rebinding / TOCTOU). For a hard
       guarantee, run the crawler with OS/network-level egress restrictions; do not
       rely on this check alone for untrusted targets.
    """
    host = get_hostname(url)
    if not host:
        return True
    if host == "localhost" or host in _METADATA_HOSTS or host.endswith(".local"):
        return True
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        # DNS failure / invalid host -> fail closed.
        log.debug("is_blocked_address: could not resolve %s - blocking", host, exc_info=True)
        return True
    for info in infos:
        sockaddr = info[4]
        ip_str = sockaddr[0]
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            return True
        # Unwrap IPv4-mapped IPv6 (e.g. ::ffff:127.0.0.1) before classifying.
        mapped = getattr(ip, "ipv4_mapped", None)
        if mapped is not None:
            ip = mapped
        if (
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_reserved
            or ip.is_multicast
            or ip.is_unspecified
        ):
            return True
    return False


def strip_tracking_params(url: str) -> str:
    """Remove UTM/tracking params and sort the query for stability."""
    try:
        p = urlparse(url)
        q = [
            (k, v)
            for (k, v) in parse_qsl(p.query, keep_blank_values=True)
            if k.lower() not in _TRACKING_PARAMS
        ]
        q.sort(key=lambda kv: kv[0].lower())
        return urlunparse(
            (p.scheme, p.netloc, p.path, p.params, urlencode(q, doseq=True), p.fragment)
        )
    except Exception:
        return url


def normalize_url(url: str) -> str:
    """
    Normalize a URL: lowercase scheme/host, drop fragment, strip tracking,
    normalize the trailing slash of the path.
    """
    try:
        url = strip_tracking_params(url.strip())
        p = urlparse(url)
        path = p.path.rstrip("/") or "/"
        return urlunparse((p.scheme.lower(), p.netloc.lower(), path, p.params, p.query, ""))
    except Exception:
        return url.strip()


def is_excluded_url(url: str, text: str = "", pattern: "Optional[re.Pattern[str]]" = None) -> bool:
    """True if the URL or anchor text matches the exclusion pattern.

    ``pattern`` defaults to the built-in exclusion regex; pass a custom compiled
    pattern (see ``compile_exclude``) to honor CrawlerConfig.exclude_patterns.
    """
    pat = pattern or _EXCLUDE_RE
    if pat.search(url):
        return True
    if text and pat.search(text):
        return True
    return False


# =============================================================================
# HASHING (dedup)
# =============================================================================


def _quiet_close(resource: object) -> None:
    """Close a resource (HTTP session / browser / DB connection), swallowing any
    error. Module-level so ``weakref.finalize`` never captures the owning object
    (which would keep it alive and defeat the finalizer)."""
    try:
        resource.close()  # type: ignore[attr-defined]
    except Exception:
        pass


def sha256_hex(s: str) -> str:
    """SHA256 hex of a UTF-8 string."""
    return hashlib.sha256(s.encode("utf-8", errors="ignore")).hexdigest()


def url_hash(url: str) -> str:
    """Level-1 dedup key (URL): sha256(normalize_url(url))."""
    return sha256_hex(normalize_url(url))


def _normalize_for_hash(text: str) -> str:
    """Normalize whitespace for a stable content hash."""
    s = (text or "").strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s


def content_hash(text: str) -> str:
    """Level-2 dedup key (content): sha256(normalize(text))."""
    return sha256_hex(_normalize_for_hash(text))


# =============================================================================
# DOMAIN BLACKLIST
# =============================================================================


def is_blacklisted_domain(url: str, blacklist: Optional[List[str]] = None) -> bool:
    """
    True if the URL host is in the blacklist (exact match or subdomain).
    E.g. "example.com" also blocks "www.example.com" and "news.example.com".
    """
    if not blacklist:
        return False
    host = get_hostname(url)
    if not host:
        return False
    blocked = {str(d).lower().strip().lstrip(".") for d in blacklist if d}
    return any(host == d or host.endswith(f".{d}") for d in blocked)


def load_blacklist_from_excel(
    excel_path: str,
    sheet_name: Optional[str] = None,
    column_name: Optional[str] = None,
) -> List[str]:
    """
    Load a list of domains from an .xlsx file.

    If column_name is None, look for a header among {domain, domains, blacklist,
    blacklisted_domain(s), blocked_domain(s)}, otherwise use the first column.
    Requires openpyxl (``pip install openpyxl``). Errors -> empty list.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        log.warning("openpyxl not installed - Excel blacklist ignored (pip install openpyxl)")
        return []

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        log.warning("error opening Excel blacklist %s: %s: %s", excel_path, type(e).__name__, e)
        return []

    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        return []

    header_lower = [str(v).strip().lower() if v is not None else "" for v in header]
    target_idx = None
    if column_name:
        wanted = column_name.strip().lower()
        target_idx = next((i for i, n in enumerate(header_lower) if n == wanted), None)
        if target_idx is None:
            log.warning("blacklist column '%s' not found - using first column", column_name)
            target_idx = 0
    else:
        candidates = {
            "domain",
            "domains",
            "blacklisted_domain",
            "blacklisted_domains",
            "blacklist",
            "blocked_domain",
            "blocked_domains",
        }
        target_idx = next((i for i, n in enumerate(header_lower) if n in candidates), 0)

    domains: List[str] = []
    seen = set()
    for row in rows:
        if row is None or target_idx >= len(row):
            continue
        value = row[target_idx]
        if value is None:
            continue
        domain = str(value).strip().lower().lstrip(".")
        if not domain:
            continue
        host = get_hostname(domain if "://" in domain else f"https://{domain}")
        domain = host or domain
        if domain not in seen:
            seen.add(domain)
            domains.append(domain)
    return domains


# =============================================================================
# HTML -> TEXT (basic fallback)
# =============================================================================


def html_to_text_basic(html: str) -> str:
    """Convert HTML to plain text via regex (fallback when trafilatura is absent)."""
    html = re.sub(r"(?is)<(script|style).*?>.*?</\1>", "", html)
    html = re.sub(r"(?is)<br\s*/?>", "\n", html)
    html = re.sub(r"(?is)</p\s*>", "\n\n", html)
    html = re.sub(r"(?is)<.*?>", "", html)
    html = (
        html.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )
    return html.strip()


# =============================================================================
# HTTP CLIENT
# =============================================================================


@dataclass
class FetchResult:
    """
    Outcome of a single fetch.

    For HTML resources ``html``/``text`` are populated. For PDFs (detected via
    Content-Type, a .pdf extension, or the %PDF- magic bytes) the raw bytes are
    returned in ``content`` and text extraction is deferred to the PDF pipeline
    — so a PDF is downloaded exactly once.
    """

    html: Optional[str] = None
    text: Optional[str] = None
    status: Optional[int] = None
    content: Optional[bytes] = None
    content_type: str = ""
    final_url: Optional[str] = None  # last hop after manual redirect following

    def __iter__(self):
        # Backward-compatible unpacking: html, text, status = client.fetch(url)
        return iter((self.html, self.text, self.status))


class HTTPClient:
    """
    HTTP client with exponential retry/backoff and text extraction via
    trafilatura (falls back to a basic HTML strip).
    """

    def __init__(self, cfg: Optional[HTTPConfig] = None):
        self.cfg = cfg or HTTPConfig()
        # SSRF + JS-rendering are mutually exclusive: the per-hop SSRF guard only
        # covers the requests path; a headless browser follows redirects and loads
        # subresources (iframes, images, scripts, XHR) that bypass it entirely.
        # Fail fast rather than offer a guard that is silently bypassed.
        if self.cfg.render_js and self.cfg.block_private_addresses:
            raise ValueError(
                "HTTPConfig(render_js=True) cannot be combined with "
                "block_private_addresses=True: the headless browser's redirects and "
                "subresource requests bypass the SSRF guard. Disable one — e.g. "
                "CrawlerTools(enforce_ssrf_guard=False) for trusted internal crawling, "
                "or render_js=False to keep the guard."
            )
        # verify=: path to the CA bundle if provided, otherwise the verify_ssl bool
        self._verify = self.cfg.ca_bundle or self.cfg.verify_ssl
        if self._verify is False:
            try:
                import urllib3

                urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            except Exception:
                log.debug("could not disable urllib3 InsecureRequestWarning", exc_info=True)
        self._session: "Optional[requests.Session]" = None
        self._browser = None
        self._browser_finalizer: "Optional[weakref.finalize]" = None
        self._finalizer: "Optional[weakref.finalize]" = None
        # Build the session now (and arm GC/exit cleanup). After a release() the
        # session is rebuilt lazily, so a tool call can free its sockets between
        # calls and the same client is reused on the next call.
        self._ensure_session()

    def _make_session(self) -> requests.Session:
        s = requests.Session()
        s.headers.update(
            {
                "User-Agent": self.cfg.user_agent,
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
            }
        )
        return s

    @property
    def session(self) -> requests.Session:
        return self._ensure_session()

    def _ensure_session(self) -> requests.Session:
        """Return the live session, lazily (re)building it after a release().

        Arms the GC/exit finalizer on the (re)built session so automatic cleanup
        keeps working across release/reuse cycles.
        """
        if self._session is None:
            self._session = self._make_session()
            self._finalizer = weakref.finalize(self, _quiet_close, self._session)
        return self._session

    def _extract_text(self, html: str) -> Optional[str]:
        """Extract main text via trafilatura, with a basic HTML-strip fallback.

        Text shorter than ``cfg.min_text_chars`` is rejected so short-but-valid
        pages (docs, changelogs, landing pages) are not silently dropped.
        """
        min_chars = self.cfg.min_text_chars
        try:
            import trafilatura  # type: ignore

            content = trafilatura.extract(
                html,
                include_comments=False,
                include_tables=False,
                no_fallback=False,
                favor_recall=True,
            )
            if content and len(content.strip()) >= min_chars:
                log.debug("  text: trafilatura -> %d chars", len(content.strip()))
                return content.strip()
            log.debug(
                "  text: trafilatura returned %s chars (<%d) -> trying basic strip",
                len(content.strip()) if content else 0,
                min_chars,
            )
        except ImportError:
            log.debug(
                "  text: trafilatura not installed -> basic HTML strip "
                "(pip install trafilatura for better extraction)"
            )
        except Exception:
            log.debug("  text: trafilatura.extract failed -> basic HTML strip", exc_info=True)
        plain = html_to_text_basic(html)
        if plain and len(plain) >= min_chars:
            log.debug("  text: basic HTML strip (fallback) -> %d chars", len(plain))
            return plain
        log.debug(
            "  text: no extractable content (<%d chars from both trafilatura and basic strip)",
            min_chars,
        )
        return None

    def _request(self, url: str, headers: Optional[dict] = None):
        """GET with **manual** redirect handling: every hop is re-validated by the
        SSRF guard (so a public host that redirects to a private address is
        blocked) and the hop count is bounded by ``max_redirects``. Returns the
        final streamed ``Response`` (caller reads+closes the body), or None if a
        hop is blocked / there are too many redirects."""
        cfg = self.cfg
        current = url
        for _ in range(cfg.max_redirects + 1):
            if cfg.block_private_addresses and is_blocked_address(current):
                log.warning("SSRF guard: refusing to fetch private/loopback address %s", current)
                return None
            resp = self.session.get(
                current,
                timeout=(cfg.timeout_connect, cfg.timeout_read),
                allow_redirects=False,
                headers=headers,
                verify=self._verify,
                stream=True,
            )
            if resp.is_redirect and resp.headers.get("Location"):
                nxt = urljoin(current, resp.headers["Location"])
                resp.close()
                current = nxt
                continue
            return resp
        log.warning("too many redirects (> %d) for %s", cfg.max_redirects, url)
        return None

    @staticmethod
    def _read_capped(resp, cap: int) -> bytes:
        """Stream the response body, stopping at ``cap`` bytes (prevents a huge or
        hostile resource from exhausting memory)."""
        total = 0
        chunks = []
        for chunk in resp.iter_content(chunk_size=65536):
            if not chunk:
                continue
            chunks.append(chunk)
            total += len(chunk)
            if total >= cap:
                log.warning("response body hit the %d-byte cap for %s - truncating", cap, resp.url)
                break
        return b"".join(chunks)[:cap]

    @staticmethod
    def _decode(body: bytes, content_type: str) -> str:
        enc = None
        if "charset=" in content_type:
            enc = content_type.split("charset=")[-1].split(";")[0].strip() or None
        return body.decode(enc or "utf-8", errors="replace")

    def fetch(
        self,
        url: str,
        extra_headers: Optional[dict] = None,
    ) -> FetchResult:
        """
        Fetch with retry. Returns a :class:`FetchResult`.

        - html:         raw HTML (None on failure or for PDFs)
        - text:         text extracted via trafilatura / fallback (None if none)
        - status:       HTTP status (None on network error)
        - content:      raw bytes for PDF resources (so they are downloaded once)
        - content_type: response Content-Type (lowercased)

        Unpacks as ``html, text, status`` for backward compatibility.

        If cfg.render_js is True, the HTML is obtained from a headless browser
        (Playwright); on browser failure/unavailability it falls back to requests.
        """
        cfg = self.cfg

        # SSRF guard (opt-in): refuse fetches to private/loopback/metadata targets.
        if cfg.block_private_addresses and is_blocked_address(url):
            log.warning("SSRF guard: refusing to fetch private/loopback address %s", url)
            return FetchResult()

        # JavaScript rendering path (opt-in).
        if cfg.render_js:
            html = self._browser_renderer().render(url)
            if html:
                return FetchResult(html=html, text=self._extract_text(html), status=200)
            # browser unavailable/failed -> fall through to requests

        for attempt in range(1, cfg.max_retries + 1):
            try:
                resp = self._request(url, extra_headers or None)
                if resp is None:
                    return FetchResult()  # blocked hop / too many redirects
                status = resp.status_code
                if status in (429, 500, 502, 503, 504):
                    resp.close()
                    raise requests.HTTPError(f"HTTP {status}")  # retryable
                if 400 <= status < 600:
                    # Permanent error (e.g. 404/403/401/410): terminal, do not retry.
                    resp.close()
                    log.info("fetch: non-retryable HTTP %s for %s - giving up", status, url)
                    return FetchResult(status=status)

                ctype = (resp.headers.get("Content-Type") or "").lower()
                final_url = resp.url  # last hop after redirects (for robots/provenance)
                looks_pdf = "application/pdf" in ctype or url.lower().split("?")[0].endswith(".pdf")
                body = self._read_capped(
                    resp, cfg.max_pdf_bytes if looks_pdf else cfg.max_html_bytes
                )
                resp.close()
                if looks_pdf or body[:5] == b"%PDF-":
                    # PDF: hand the bytes straight to the PDF pipeline (no re-download).
                    return FetchResult(
                        status=status, content=body, content_type=ctype, final_url=final_url
                    )

                html = self._decode(body, ctype)
                return FetchResult(
                    html=html,
                    text=self._extract_text(html),
                    status=status,
                    content_type=ctype,
                    final_url=final_url,
                )

            except Exception as e:
                if attempt < cfg.max_retries:
                    log.debug(
                        "fetch attempt %d/%d for %s failed: %s", attempt, cfg.max_retries, url, e
                    )
                    time.sleep(cfg.backoff_base_sec * (2 ** (attempt - 1)))
                else:
                    log.warning(
                        "fetch failed for %s after %d attempts: %s: %s",
                        url,
                        cfg.max_retries,
                        type(e).__name__,
                        e,
                    )
                    return FetchResult()

        return FetchResult()

    def fetch_bytes(self, url: str) -> "tuple[Optional[bytes], str, Optional[int]]":
        """
        Raw GET for a binary asset (e.g. an image). Honors the SSRF guard and the
        SSL/verify configuration. Returns ``(bytes, content_type, status)``;
        ``bytes`` is None on block/failure or a >=400 response.
        """
        cfg = self.cfg
        try:
            resp = self._request(url, None)  # per-hop SSRF validation inside
            if resp is None:
                return None, "", None
            ctype = (resp.headers.get("Content-Type") or "").lower()
            if resp.status_code >= 400:
                resp.close()
                return None, ctype, resp.status_code
            body = self._read_capped(resp, cfg.max_asset_bytes)
            status = resp.status_code
            resp.close()
            return body, ctype, status
        except Exception as e:
            log.debug("fetch_bytes failed for %s: %s", url, e)
            return None, "", None

    def get_text(self, url: str) -> Optional[str]:
        """
        Fetch a URL and return the raw response text (no extraction), honoring
        the SSL/verify configuration and the SSRF guard. Used for robots.txt.
        None on any failure.
        """
        try:
            resp = self._request(url, None)
            if resp is None or resp.status_code >= 400:
                if resp is not None:
                    resp.close()
                return None
            body = self._read_capped(resp, self.cfg.max_html_bytes)
            ctype = (resp.headers.get("Content-Type") or "").lower()
            resp.close()
            return self._decode(body, ctype)
        except Exception as e:
            log.debug("get_text failed for %s: %s", url, e)
            return None

    def release(self) -> None:
        """Free OS resources (sockets + browser) but stay reusable.

        The session is rebuilt lazily on the next request, so a tool call can
        release everything it opened **at the end of the call** without discarding
        the client (config, robots/rate caches live elsewhere). The GC/exit
        finalizer is disarmed here and re-armed when the session is rebuilt.
        """
        if self._browser is not None:
            try:
                self._browser.close()
            except Exception:
                log.debug("failed closing browser renderer", exc_info=True)
            self._browser = None
            if self._browser_finalizer is not None:
                self._browser_finalizer.detach()
                self._browser_finalizer = None
        if self._session is not None:
            if self._finalizer is not None:
                self._finalizer.detach()
            _quiet_close(self._session)
            self._session = None

    # close() is an alias: release() already frees every OS resource, and the
    # client stays reusable (lazy session) — harmless after an explicit close().
    close = release

    def __enter__(self) -> "HTTPClient":
        return self

    def __exit__(self, *exc) -> bool:
        self.close()
        return False

    def _browser_renderer(self):
        if self._browser is None:
            from .browser import BrowserRenderer

            cfg = self.cfg
            self._browser = BrowserRenderer(
                user_agent=cfg.user_agent,
                headless=cfg.browser_headless,
                wait_until=cfg.browser_wait_until,
                timeout_ms=cfg.browser_timeout_ms,
            )
            # Close the browser subprocess too on GC/exit if close() is never called.
            self._browser_finalizer = weakref.finalize(self, _quiet_close, self._browser)
        return self._browser


# =============================================================================
# ROBOTS.TXT
# =============================================================================


class RobotsChecker:
    """
    Thread-safe robots.txt gate. Fetches each host's robots.txt once (honoring
    the SSL config via HTTPClient) and answers can_fetch for the configured
    User-Agent. A missing/unreachable/unparseable robots.txt means "allow"
    (standard convention).
    """

    def __init__(self, http: HTTPClient, user_agent: str):
        self._http = http
        self._ua = user_agent or "*"
        self._cache: Dict[str, Optional[RobotFileParser]] = {}
        self._lock = threading.Lock()
        # Per-host locks so two threads never fetch the same robots.txt twice
        # (and never block on *other* hosts while one fetch is in flight).
        self._host_locks: Dict[str, threading.Lock] = {}

    def allowed(self, url: str) -> bool:
        try:
            p = urlparse(url)
        except Exception:
            return True
        host = (p.netloc or "").lower()
        if not host:
            return True
        rp = self._get(p.scheme or "https", host)
        if rp is None:
            return True
        try:
            return rp.can_fetch(self._ua, url)
        except Exception:
            log.debug("robots can_fetch errored for %s - allowing", url, exc_info=True)
            return True

    def crawl_delay(self, url: str) -> Optional[float]:
        """robots.txt Crawl-delay (seconds) for the configured UA, or None."""
        try:
            p = urlparse(url)
        except Exception:
            return None
        host = (p.netloc or "").lower()
        if not host:
            return None
        rp = self._get(p.scheme or "https", host)
        if rp is None:
            return None
        try:
            delay = rp.crawl_delay(self._ua)
            return float(delay) if delay is not None else None
        except Exception:
            return None

    def _host_lock(self, host: str) -> threading.Lock:
        with self._lock:
            lk = self._host_locks.get(host)
            if lk is None:
                lk = threading.Lock()
                self._host_locks[host] = lk
            return lk

    def _get(self, scheme: str, host: str) -> Optional[RobotFileParser]:
        with self._lock:
            if host in self._cache:
                return self._cache[host]
        # Serialize per-host so the robots.txt for this host is fetched once.
        with self._host_lock(host):
            with self._lock:
                if host in self._cache:  # double-check after acquiring the host lock
                    return self._cache[host]
            robots_url = urljoin(f"{scheme}://{host}", "/robots.txt")
            rp: Optional[RobotFileParser] = None
            text = self._http.get_text(robots_url)
            if text:
                rp = RobotFileParser()
                try:
                    rp.parse(text.splitlines())
                except Exception:
                    log.debug(
                        "failed to parse robots.txt at %s - allowing", robots_url, exc_info=True
                    )
                    rp = None
            with self._lock:
                self._cache[host] = rp
            return rp
